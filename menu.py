from tkinter import * 
from tkinter import ttk
from openpyxl import load_workbook 
from tkinter import messagebox
import pandas as pd

class menuWin(Frame):
    def __init__(self, root=None):
        super().__init__(root, width=800, height=440)
        self.root = root
        self.pack()

        self.widgetInput()
        self.widgetUpBtn()
        self.widgetTable()
        self.toUpdateTable()
        self.widgetBusqueda()
        self.widgetDownBtn()
        
    def widgetInput(self):
        self.lblFrameInput = LabelFrame(self, text="Datos")
        self.lblFrameInput.grid(column=0, row=0)
        self.lblFrameInput.config(width="800", height="200", padx=5, pady=5)
        self.lblFrameInput.place(x=345, y=45, anchor=CENTER)
        #Nombre
        self.name = Entry(self.lblFrameInput)
        self.name.grid(column=1, row=0)
        self.name.config(width="30", justify="center")
        self.lblName = Label(self.lblFrameInput, text="Nombre")
        self.lblName.grid(column=0, row=0, )
        self.lblName.config(width="10", justify="center")
        #Apellido
        self.surname = Entry(self.lblFrameInput)
        self.surname.grid(column=3, row=0)
        self.surname.config(width="30", justify="center")
        self.lblSurname = Label(self.lblFrameInput, text="Apellido")
        self.lblSurname.grid(column=2, row=0)
        self.lblSurname.config(width="10", justify="center")
        #Listado de jornadas
        self.jornada = StringVar()
        self.jornada.set("Jornada")
        self.combxJornada = OptionMenu(self.lblFrameInput, self.jornada, "Diurna","Nocturna")
        self.combxJornada.grid(column=2, row=2)
        self.combxJornada.config(width="8", justify="center")
        #Listado de carreras
        self.carrera = StringVar(self.lblFrameInput)
        self.carrera.set("Carrera")
        self.menuCarrera = OptionMenu(self.lblFrameInput, self.carrera,"Administración de Empresas", "Arquitectura", "Comunicación Social", "Contaduría Pública", "Derecho", "Ingeniería de Sistemas", "Ingeniería de Software", "Ingeniería Industrial", "Psicología", "Auxiliar en Enfermería", "Auxiliar Clínica Veterinaria y Cuidado de Mascotas", "Animación Digital 2D y 3D", "Diseño de Modas", "Auxiliar en Comunicación Gráfica y Publicitaria", "Operaciones de Software y Redes de Cómputo", "Auxiliar Administrativo", "Judicial y Criminalística", "Seguridad Ocupacional", "Auxiliar Contable y Financiero", "Auxiliar en Recursos Humanos", "Cocina Nacional e Internacional", "Especialización en Derecho Administrativo y Contractual", "Especialización en Derecho Penal y Criminalística", "Especialización en Gerencia Financiera", "Especialización en Gerencia de Empresas", "Especialización en Gerencia del Talento Humano")
        self.menuCarrera.grid(column=1, row=2)
        self.menuCarrera.config(width="42", justify="center")
        #Listado semestres
        self.semestre = StringVar()
        self.semestre.set("Semestre")
        self.comboSemestre = OptionMenu(self.lblFrameInput, self.semestre, "1","2","3","4","5","6","7","8","9","10")
        self.comboSemestre.grid(column=3, row=2)
        self.comboSemestre.config(width="7", justify="center")

    def widgetUpBtn(self):
        #Frame Btn arriba
        self.btnSave = Button(self, text="Guardar", command=self.saveData)
        self.btnSave.place(x=735, y=50, anchor=CENTER)
        self.btnSave.config(width="10", justify="center", bg="#856ff8", fg="white")      

    def widgetTable(self):
        #Frame tabla
        self.lblFrameTable = LabelFrame(self, text="Tabla")
        self.lblFrameTable.grid(column=0, row=0)
        self.lblFrameTable.config(width="770", height="300", padx=5, pady=5)
        self.lblFrameTable.place(x=400, y=230, anchor=CENTER)

        self.tabla = ttk.Treeview(self.lblFrameTable)
        self.tabla.grid(column=0, row=0)
        ladox = Scrollbar(self.lblFrameTable, orient=HORIZONTAL, command=self.tabla.xview)
        ladoy = Scrollbar(self.lblFrameTable, orient=VERTICAL, command=self.tabla.yview)
        self.tabla.configure(xscrollcommand=ladox.set, yscrollcommand=ladoy.set)
        ladox.grid(column=0, row=1, sticky="ew")
        ladoy.grid(column=1, row=0, sticky="ns")
        
        self.tabla["columns"] = ("Nombre", "Apellido", "Jornada", "Carrera", "Semestre")
        self.tabla.column("#0", width=0, stretch=NO)
        
        self.tabla.column("Nombre", anchor=CENTER, width=150)
        self.tabla.column("Apellido", anchor=CENTER, width=150)
        self.tabla.column("Jornada", anchor=CENTER, width=90)
        self.tabla.column("Carrera", anchor=CENTER, width=280)
        self.tabla.column("Semestre", anchor=CENTER, width=80)    

        self.tabla.heading("#0", text="", anchor=CENTER)
        self.tabla.heading("Nombre", text="Nombre", anchor=CENTER)
        self.tabla.heading("Apellido", text="Apellido", anchor=CENTER)
        self.tabla.heading("Jornada", text="Jornada", anchor=CENTER)
        self.tabla.heading("Carrera", text="Carrera", anchor=CENTER)
        self.tabla.heading("Semestre", text="Semestre", anchor=CENTER)

    def widgetBusqueda(self):
        #LabelFrame Busqueda
        self.frameSearch = Frame(self)
        self.frameSearch.grid(column=0, row=0)
        self.frameSearch.config(width="770", height="300", padx=5, pady=5)
        self.frameSearch.place(x=220, y=400, anchor=CENTER)
        #Entry Busqueda
        self.entrySearch = Entry(self.frameSearch)
        self.entrySearch.grid(column=1, row=0)
        self.entrySearch.config(width="50", justify="center")
        #Boton Busqueda
        self.btnSearch = Button(self.frameSearch, text="Buscar",command=self.Buscar)
        self.btnSearch.grid(column=2, row=0)
        self.btnSearch.config(width="10", justify="center")

    def widgetDownBtn(self):
        self.frameBtn = Frame(self)
        self.frameBtn.config(width="200", height="200", padx=5, pady=5)
        self.frameBtn.place(x=700, y=400, anchor=CENTER)
        #Boton Actualizar
        self.btnUpdate = Button(self.frameBtn, text="Actualizar", command=self.toUpdateTable)
        self.btnUpdate.grid(column=0, row=0)
        self.btnUpdate.config(width="10", justify="center")
        #Boton Eliminar
        self.btnDelete = Button(self.frameBtn, text="Eliminar", command=self.deleteData)
        self.btnDelete.grid(column=1, row=0)
        self.btnDelete.config(width="10", justify="center", bg="#F06C50", fg="white")

    def Buscar(self):
        if self.entrySearch.get() == "":
            messagebox.showinfo("Error", "Escriba algo en el campo de busqueda")
        else:
            self.tabla.delete(*self.tabla.get_children())
            self.wb = load_workbook(filename = 'BD.xlsx')
            self.ws = self.wb.active 
            for row in self.ws.iter_rows(min_row=2, max_col=5, max_row=self.ws.max_row):
                if self.entrySearch.get() in row[0].value:
                    self.tabla.insert("", END, text="", values=(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value))

    def toUpdateTable(self):
        self.tabla.delete(*self.tabla.get_children())
        self.wb = load_workbook(filename = 'BD.xlsx')
        self.ws = self.wb.active
        for row in self.ws.iter_rows(min_row=2, max_col=5, max_row=self.ws.max_row):
            self.tabla.insert("", END, text="", values=(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value))

    def saveData(self):
        if self.name.get() == "" or self.surname.get() == "" or self.jornada.get() == "Jornada" or self.carrera.get() == "Carrera" or self.semestre.get() == "Semestre":
            messagebox.showwarning("Advertencia", "Debe ingresar todos los datos")
        else:
            name = self.name.get()  
            surname = self.surname.get()
            jornada = self.jornada.get()
            carrera = self.carrera.get()
            semestre = self.semestre.get()
            df = pd.read_excel("BD.xlsx") 
            df.loc[len(df.index)] = [name, surname, jornada, carrera, semestre]
            df.to_excel("BD.xlsx", index=False)  
            self.name.delete(0, END) 
            self.surname.delete(0, END) 
            self.jornada.set("Jornada")
            self.carrera.set("Carrera")
            self.semestre.set("Semestre")
            self.tabla.delete(*self.tabla.get_children())
            self.toUpdateTable()
            messagebox.showinfo("Información", "Datos guardados correctamente") 

    def deleteData(self):    
        if not self.tabla.selection():
            messagebox.showwarning("Advertencia", "Por favor seleccione un registro")
        else:
            item = self.tabla.selection()[0]
            self.wb = load_workbook(filename = 'BD.xlsx')
            self.ws = self.wb.active
            for row in self.ws.iter_rows(min_row=2, max_col=5, max_row=self.ws.max_row):
                if self.tabla.item(item, "values")[0] == row[0].value:
                    self.ws.delete_rows(row[0].row)
                    self.wb.save('BD.xlsx')
                    self.tabla.delete(item)
                    messagebox.showinfo("Información", "Registro eliminado correctamente")
                    break

if __name__ == '__main__':
    root = Tk()
    app = menuWin(root)
    root.title("Control de datos")
    root.iconbitmap("icon.ico") 
    root.mainloop()